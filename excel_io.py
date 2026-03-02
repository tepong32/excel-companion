"""
excel_io.py
Excel file I/O and user preferences — no Tkinter dependency.
"""

import json
import os
import warnings

from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# ---------------------------------------------------------------------------
# Prefs path
# FIX: Prefs are now stored in a proper per-user config directory instead of
#      next to the Excel file (which pollutes shared/cloud-synced folders).
# ---------------------------------------------------------------------------

def _get_prefs_dir() -> str:
    """Return (and create if needed) an OS-appropriate app config directory."""
    try:
        import platformdirs
        d = platformdirs.user_config_dir("tEppy_DataEntry")
    except ImportError:
        # Graceful fallback if platformdirs isn't installed
        d = os.path.join(os.path.expanduser("~"), ".tEppy_DataEntry")
    os.makedirs(d, exist_ok=True)
    return d


def get_prefs_path(filepath: str) -> str | None:
    """
    Return a stable prefs path derived from the Excel filepath.
    The filename is a sanitised version of the full path so it's unique
    even if two files share the same basename.
    """
    if not filepath:
        return None
    # Replace path separators and colons with underscores, keep it short
    safe = filepath.replace(os.sep, "_").replace(":", "_").replace("/", "_")
    # Truncate from the right so we keep the meaningful end of the path
    safe = safe[-120:] + ".prefs.json"
    return os.path.join(_get_prefs_dir(), safe)


# ---------------------------------------------------------------------------
# Universal Excel loader
# ---------------------------------------------------------------------------

def load_any_excel(path: str, warn_callback=None):
    """
    Load any supported spreadsheet format and return an openpyxl Workbook.
    Non-.xlsx formats are converted to an in-memory .xlsx representation.

    warn_callback: optional callable(message: str) for UI warnings.
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            with warnings.catch_warnings(record=True):
                warnings.simplefilter("always")
                wb = load_workbook(path, data_only=True)

            # Warn about embedded image formats that may break saving
            unsupported_formats = (".wmf", ".emf", ".tiff", ".bmp")
            for ws in wb.worksheets:
                for image in getattr(ws, "_images", []):
                    img_path = str(getattr(image, "path", "")).lower()
                    if any(fmt in img_path for fmt in unsupported_formats):
                        if warn_callback:
                            warn_callback("⚠️ Workbook contains embedded images (WMF/EMF/TIFF/BMP).")
                        break
            return wb

        elif ext == ".xlsb":
            df = pd.read_excel(path, engine="pyxlsb")
        elif ext == ".xls":
            df = pd.read_excel(path, engine="xlrd")
        elif ext == ".ods":
            df = pd.read_excel(path, engine="odf")
        else:
            raise ValueError(f"Unsupported file format: {ext!r}")

        # Convert DataFrame → openpyxl Workbook
        wb = Workbook()
        ws = wb.active
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
        return wb

    except Exception as e:
        raise RuntimeError(f"Failed to load file ({ext}): {e}") from e


# ---------------------------------------------------------------------------
# Preferences (read / write)
# ---------------------------------------------------------------------------

def load_user_prefs(filepath: str) -> dict:
    """Return saved prefs dict, or {} if none exist."""
    prefs_path = get_prefs_path(filepath)
    if not prefs_path or not os.path.exists(prefs_path):
        return {}
    try:
        with open(prefs_path, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


def save_user_prefs(filepath: str, prefs: dict) -> bool:
    """Persist prefs dict. Returns True on success."""
    prefs_path = get_prefs_path(filepath)
    if not prefs_path:
        return False
    try:
        # Merge with existing prefs so unrelated keys survive
        existing = {}
        if os.path.exists(prefs_path):
            try:
                with open(prefs_path, "r", encoding="utf-8") as f:
                    existing = json.load(f) or {}
            except Exception:
                pass
        existing.update(prefs)
        with open(prefs_path, "w", encoding="utf-8") as f:
            json.dump(existing, f, indent=2)
        return True
    except Exception:
        return False
