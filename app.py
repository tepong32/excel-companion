"""
app.py
DynamicExcelApp — the main application class.

All fixes applied (see CHANGELOG for details):
  - Cross-platform help file opening
  - Shadow desync after row deletion (via ShadowStore.delete_row)
  - Duplicate try/except no-op in add_row_from_inputs
  - Ctrl+N now correctly clears fields (not new file); new file moved to Ctrl+Shift+N
  - Bare except clauses replaced with logged exceptions (set DEBUG=1 env var)
  - Duplicate tooltip system removed (ToolTip class only)
  - _flush_shadow_to_workbook stub removed
  - Ctrl+D (duplicate row) implemented
  - Ctrl+Shift+I (insert blank row) implemented
  - ID detection uses regex (validation.py) to avoid false positives
  - Prefs now stored in OS user-config dir via excel_io.py
"""

import os
import re
import sys
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkbootstrap import Window, Style
from openpyxl import Workbook

from validation import (
    infer_validation_rules,
    format_value_for_display,
    detect_precision_mismatch,
    has_excess_precision,
    is_numeric,
    normalize_numeric,
    try_parse_date,
    round_half_up,
    EMAIL_RE,
)
from excel_io import load_any_excel, load_user_prefs, save_user_prefs
from widgets import ToolTip, ShadowStore

APP_TITLE = "tEppy's Data Entry (Excel Companion with validation)"

# Set DEBUG=1 environment variable to see suppressed exceptions in the console.
DEBUG = os.getenv("APP_DEBUG", "0") == "1"


def _log(label: str, exc: Exception):
    """Log exceptions when DEBUG is active instead of silently swallowing them."""
    if DEBUG:
        print(f"[DEBUG:{label}] {type(exc).__name__}: {exc}")


def resource_path(relative_path: str) -> str:
    """Resolve path for both normal runs and PyInstaller bundles."""
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


def open_file_cross_platform(path: str):
    """
    FIX: os.startfile() is Windows-only and silently did nothing on macOS/Linux.
    This helper works on all three platforms.
    """
    if os.name == "nt":
        os.startfile(path)
    elif sys.platform == "darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


# ---------------------------------------------------------------------------
# Main Application
# ---------------------------------------------------------------------------

class DynamicExcelApp:
    def __init__(self, root: Window):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1100x720")

        # --- Core state ---
        self.workbook = None
        self.filepath = None
        self.active_sheet_name = None
        self.current_sheet = None
        self.headers = []
        self.input_entries = []
        self.unsaved_changes = False
        self.original_editing_values = {}
        self.mode = "add"
        self.editing_item = None
        self.selected_item = None

        # Validation
        self.validation_rules = []
        self.current_rules = {}
        self.input_order = []

        # FIX: replaced bare shadow_values dict with ShadowStore which handles
        #      re-indexing after deletions automatically.
        self.shadow = ShadowStore()

        # Treeview row cache (for filtering)
        self.all_rows = []
        self.filter_entries = []
        self.column_ids = []

        # Visual helpers
        self.style = ttk.Style()
        self._generate_rounded_icon()

        # UI construction
        self._create_menu()
        self._create_toolbar()
        self._create_top_frame()
        self._create_bottom_frame()
        self._create_statusbar()

        self._bind_events()
        self._bind_shortcuts()

        self._prompt_open_file_on_startup()

    # -----------------------------------------------------------------------
    # Icon generation
    # -----------------------------------------------------------------------

    def _generate_rounded_icon(self):
        try:
            self.rounded_icon = tk.PhotoImage(width=12, height=12)
            data = []
            for y in range(12):
                row = []
                for x in range(12):
                    if 1 <= x <= 10 and 1 <= y <= 10 and y >= abs(x - 6) + 2:
                        row.append("#d9534f")
                    else:
                        row.append("None")
                data.append("{" + " ".join(row) + "}")
            self.rounded_icon.put(" ".join(data))
        except Exception as e:
            _log("rounded_icon", e)
            self.rounded_icon = None

    # -----------------------------------------------------------------------
    # Status bar helpers
    # -----------------------------------------------------------------------

    def _set_status(self, message: str):
        """Lightweight status (hover events)."""
        if not message:
            self.status_var.set("Ready")
            try:
                self.status_label.configure(background="#f8f9fa", foreground="#333333")
            except Exception as e:
                _log("_set_status/clear", e)
            return
        try:
            self.status_var.set(message)
            self.status_label.configure(background="#fff3cd", foreground="#664d03")
        except Exception as e:
            _log("_set_status", e)
            self.status_var.set(message)

    def _update_status(self, message: str, level: str = "info", duration: int = 5000):
        colors = {
            "info": ("#f8f9fa", "#333333"),
            "success": ("#d1e7dd", "#0f5132"),
            "warning": ("#fff3cd", "#664d03"),
            "error": ("#f8d7da", "#842029"),
        }
        bg, fg = colors.get(level, colors["info"])
        try:
            self.status_var.set(message)
            self.status_label.configure(background=bg, foreground=fg)
            if duration > 0:
                self.root.after(duration, self._fade_status)
        except Exception as e:
            _log("_update_status", e)

    def _fade_status(self):
        try:
            self.status_label.configure(background="#f8f9fa", foreground="#333333")
            self.status_var.set("Ready")
        except Exception as e:
            _log("_fade_status", e)

    # -----------------------------------------------------------------------
    # Event bindings
    # -----------------------------------------------------------------------

    def _bind_events(self):
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        self.tree.bind("<Motion>", self._on_tree_hover)
        self.tree.bind("<<TreeviewSelect>>", self._on_row_select)
        self.tree.bind("<Delete>", lambda e: self.delete_selected_row())

    def _bind_shortcuts(self):
        self.root.bind("<Control-o>", lambda e: self.open_file())
        self.root.bind("<Control-s>", lambda e: self.save_file())
        self.root.bind("<Control-S>", lambda e: self.save_file_as())

        # FIX: Ctrl+N now clears input fields (matches README).
        #      New file moved to Ctrl+Shift+N.
        self.root.bind("<Control-n>", lambda e: self.reset_to_add_mode())
        self.root.bind("<Control-N>", lambda e: self.new_file())

        self.root.bind("<F2>", lambda e: self.edit_selected_row())
        self.root.bind("<Escape>", lambda e: self.reset_to_add_mode())

        # FIX: Ctrl+D (duplicate row) — was documented but never implemented.
        self.root.bind("<Control-d>", lambda e: self.duplicate_selected_row())

        # FIX: Ctrl+Shift+I (insert blank row) — was documented but never implemented.
        self.root.bind("<Control-I>", lambda e: self.insert_blank_row())

        # Ctrl+Shift+D — delete row (keep existing behaviour)
        self.root.bind("<Control-D>", lambda e: self.delete_selected_row())

    def _on_row_select(self, event=None):
        try:
            selection = self.tree.selection()
            if not selection:
                return
            item_id = selection[0]
            self.selected_item = item_id
            row_index = int(self.tree.index(item_id)) + 1
            total = len(self.tree.get_children())
            self._set_status(f"Row {row_index} of {total}")
        except Exception as e:
            _log("_on_row_select", e)

    # -----------------------------------------------------------------------
    # UI construction
    # -----------------------------------------------------------------------

    def _create_menu(self):
        menubar = tk.Menu(self.root)

        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="New          Ctrl+Shift+N", command=self.new_file)
        file_menu.add_command(label="Open...      Ctrl+O", command=self.open_file)
        file_menu.add_command(label="Save         Ctrl+S", command=self.save_file)
        file_menu.add_command(label="Save As...   Ctrl+Shift+S", command=self.save_file_as)
        file_menu.add_separator()
        file_menu.add_command(label="Exit         Ctrl+Q", command=self.on_close)
        menubar.add_cascade(label="File", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Help & Instructions", command=self._show_help)
        help_menu.add_separator()
        help_menu.add_command(label="About", command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)
        self.root.bind("<Control-q>", lambda e: self.on_close())

    def _create_toolbar(self):
        toolbar = ttk.Frame(self.root, padding=(8, 6))
        toolbar.pack(side=tk.TOP, fill=tk.X)

        delete_btn = ttk.Button(toolbar, text="🗑 Delete Row",
                                command=self.delete_selected_row, style="danger.TButton")
        delete_btn.pack(side=tk.LEFT, padx=(0, 4))
        ToolTip(delete_btn, "Delete selected row (Ctrl+Shift+D)")

        dup_btn = ttk.Button(toolbar, text="⧉ Duplicate",
                             command=self.duplicate_selected_row, style="secondary.TButton")
        dup_btn.pack(side=tk.LEFT, padx=(0, 10))
        ToolTip(dup_btn, "Copy row into input fields (Ctrl+D)")

        self.add_button = ttk.Button(toolbar, text="➕ Add Row",
                                     command=self.add_row_from_inputs, style="success.TButton")
        self.add_button.pack(side=tk.LEFT, padx=(0, 14))

        ttk.Label(toolbar, text="Sheet:", bootstyle="secondary").pack(side=tk.LEFT, padx=(8, 4))
        self.sheet_combo = ttk.Combobox(toolbar, state="readonly", width=28)
        self.sheet_combo.pack(side=tk.LEFT, padx=(0, 8))
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_change)

        ttk.Label(toolbar, text="").pack(side=tk.LEFT, expand=True)  # spacer

        right_grp = ttk.Frame(toolbar, padding=(8, 4))
        right_grp.pack(side=tk.RIGHT)
        right_grp.config(relief=tk.GROOVE, borderwidth=1)

        self.auto_save_var = tk.BooleanVar(value=False)
        auto_save_chk = ttk.Checkbutton(right_grp, text="Auto-Save",
                                         variable=self.auto_save_var,
                                         style="primary.TCheckbutton")
        auto_save_chk.pack(side=tk.LEFT, padx=(0, 8))

        ttk.Separator(right_grp, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=(4, 10))

        ttk.Label(right_grp, text="Theme:", bootstyle="secondary").pack(side=tk.LEFT, padx=(4, 4))
        self.theme_combo = ttk.Combobox(right_grp, values=Style().theme_names(),
                                         state="readonly", width=15)
        self.theme_combo.set(Style().theme_use())
        self.theme_combo.bind("<<ComboboxSelected>>", self.on_theme_change)
        self.theme_combo.pack(side=tk.LEFT, padx=(0, 6))

        help_btn = ttk.Button(right_grp, text="❓ Help", command=self._show_help,
                              style="info.TButton", width=8)
        help_btn.pack(side=tk.LEFT, padx=(10, 0))
        ToolTip(help_btn, "View usage instructions")
        self._add_hover_effect(help_btn)

    def _create_statusbar(self):
        self.status_var = tk.StringVar(value="Ready")
        self.status_label = ttk.Label(self.root, textvariable=self.status_var,
                                       anchor="e", padding=(6, 2), bootstyle="secondary")
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)

    def _create_top_frame(self):
        self.top_frame = ttk.Frame(self.root)
        self.top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(6, 0))

        self.input_canvas = tk.Canvas(self.top_frame, height=140)
        self.input_canvas.pack(side=tk.TOP, fill=tk.X, expand=True)

        self.input_scrollbar = ttk.Scrollbar(self.top_frame, orient="horizontal",
                                              command=self.input_canvas.xview)
        self.input_scrollbar.pack(side=tk.TOP, fill=tk.X)
        self.input_canvas.configure(xscrollcommand=self.input_scrollbar.set)

        self.inputs_inner = ttk.Frame(self.input_canvas)
        self.input_canvas.create_window((0, 0), window=self.inputs_inner, anchor="nw")
        self.inputs_inner.bind(
            "<Configure>",
            lambda e: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all"))
        )

    def _create_bottom_frame(self):
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.filter_frame = ttk.Frame(bottom_frame)
        self.filter_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))

        tree_frame = ttk.Frame(bottom_frame)
        tree_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(tree_frame, show="headings")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.LEFT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.configure(xscrollcommand=hsb.set)

        self.tree.bind("<Configure>", lambda e: self._adjust_filter_widths())
        self.tree.bind("<ButtonRelease-1>", lambda e: self._adjust_filter_widths())

    # -----------------------------------------------------------------------
    # Hover: show raw vs rounded in status bar
    # -----------------------------------------------------------------------

    def _on_tree_hover(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            self._set_status("")
            return

        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            self._set_status("")
            return

        try:
            col_index = int(col_id.replace("#", ""))
            row_index = int(self.tree.index(row_id)) + 1
        except Exception as e:
            _log("_on_tree_hover/parse", e)
            self._set_status("")
            return

        entry = self.shadow.get(self.current_sheet, row_index, col_index)
        if entry and entry.get("rounded_flag"):
            raw = entry.get("raw")
            displayed = self.tree.set(row_id, col_index - 1)
            self._set_status(f"Rounded: original '{raw}' → displayed '{displayed}'")
        else:
            self._set_status("")

    # -----------------------------------------------------------------------
    # Filters
    # -----------------------------------------------------------------------

    def _create_filter_row(self):
        for w in self.filter_frame.winfo_children():
            w.destroy()
        self.filter_entries.clear()
        self.column_ids = list(self.tree["columns"])
        if not self.column_ids:
            return
        for col_id in self.column_ids:
            entry = ttk.Entry(self.filter_frame)
            entry.pack(side=tk.LEFT, padx=1, fill=tk.X, expand=True)
            entry.bind("<KeyRelease>", lambda e: self._apply_filters())
            self.filter_entries.append(entry)
        self.root.after(100, self._adjust_filter_widths)

    def _adjust_filter_widths(self):
        if not self.filter_entries or not self.column_ids:
            return
        for i, col_id in enumerate(self.column_ids):
            try:
                width = int(self.tree.column(col_id, "width"))
            except Exception:
                width = 100
            self.filter_entries[i].config(width=max(8, width // 10))

    def _apply_filters(self):
        if not self.all_rows:
            return
        filters = [f.get().strip().lower() for f in self.filter_entries]
        if all(f == "" for f in filters):
            self._reload_tree_from_cache()
            return
        filtered = [
            row for row in self.all_rows
            if all((f in str(row[i]).lower() if f else True) for i, f in enumerate(filters))
        ]
        self._reload_tree_from_cache(filtered)
        self._update_status(
            f"Filtered: {len(filtered)} of {len(self.all_rows)} rows shown", "info", 3000
        )

    def _reload_tree_from_cache(self, rows=None):
        self.tree.delete(*self.tree.get_children())
        display_rows = rows if rows is not None else self.all_rows
        for row in display_rows:
            padded = list(row) + [""] * (len(self.headers) - len(row))
            self.tree.insert("", tk.END, values=padded)

    def _clear_treeview(self):
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()

    # -----------------------------------------------------------------------
    # File operations
    # -----------------------------------------------------------------------

    def _prompt_open_file_on_startup(self):
        answer = messagebox.askyesno("Open file", "Is your template ready for loading?")
        if answer:
            self.open_file()
        else:
            self.status_var.set("Ready. Use File → Open or Ctrl+O to open a spreadsheet.")

    def open_file(self):
        filetypes = [
            ("All supported files", "*.xlsx *.xlsm *.xlsb *.xls *.ods"),
            ("Excel files", "*.xlsx;*.xlsm;*.xlsb;*.xls"),
        ]
        path = filedialog.askopenfilename(title="Open spreadsheet", filetypes=filetypes)
        if not path:
            return
        try:
            wb = load_any_excel(path, warn_callback=lambda m: messagebox.showwarning("Warning", m))
        except RuntimeError as e:
            messagebox.showerror("Error", f"Failed to open file:\n{e}")
            return

        self.workbook = wb
        self.filepath = path
        try:
            self.active_sheet_name = self.workbook.active.title
        except Exception:
            self.active_sheet_name = (self.workbook.sheetnames[0]
                                       if self.workbook.sheetnames else None)

        self._populate_sheet_selector()
        self._load_active_sheet()
        self.unsaved_changes = False

        ext = os.path.splitext(path)[1].lower()
        if ext not in (".xlsx", ".xlsm"):
            self._update_status(f"Opened {os.path.basename(path)} (converted to in-memory .xlsx)")
            messagebox.showinfo(
                "Format Notice",
                "This file was opened from a non-.xlsx format.\n"
                "It will be saved as .xlsx when you save changes."
            )
        else:
            self._update_status(f"Opened: {os.path.basename(path)}")

    def new_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        self.workbook = wb
        self.filepath = None
        self.active_sheet_name = ws.title
        self._populate_sheet_selector()
        self._load_active_sheet()
        self._update_status("New workbook created.", "success")

    def _populate_sheet_selector(self):
        if not self.workbook:
            self.sheet_combo["values"] = []
            return
        names = self.workbook.sheetnames
        self.sheet_combo["values"] = names
        if self.active_sheet_name in names:
            self.sheet_combo.set(self.active_sheet_name)
        else:
            self.sheet_combo.set(names[0] if names else "")
            self.active_sheet_name = names[0] if names else None

    def on_sheet_change(self, event=None):
        if not self.workbook:
            return
        new_sheet = self.sheet_combo.get()
        if new_sheet == self.active_sheet_name:
            return
        if self.unsaved_changes:
            res = messagebox.askyesnocancel(
                "Unsaved changes", "You have unsaved changes. Save before switching sheets?"
            )
            if res is None:
                self.sheet_combo.set(self.active_sheet_name)
                return
            if res:
                if not self.save_file():
                    self.sheet_combo.set(self.active_sheet_name)
                    return
        self.active_sheet_name = new_sheet
        self._load_active_sheet()

    def _load_active_sheet(self):
        if not self.workbook or not self.active_sheet_name:
            return
        sheet = self.workbook[self.active_sheet_name]
        self.current_sheet = self.active_sheet_name

        # Detect headers in first 5 rows
        headers = []
        header_row_idx = None
        for r in sheet.iter_rows(min_row=1, max_row=5):
            values = [cell.value for cell in r]
            if any(v is not None and str(v).strip() != "" for v in values):
                header_row_idx = r[0].row
                headers = [
                    str(c.value).strip() if c.value is not None and str(c.value).strip() else None
                    for c in r
                ]
                break

        if not headers:
            max_col = sheet.max_column or 1
            headers = [None] * max_col
            header_row_idx = 1

        self.headers = [h if h else f"Column {i + 1}" for i, h in enumerate(headers)]
        self.validation_rules = infer_validation_rules(self.headers, tk)
        self.current_rules = {rule["name"]: rule for rule in self.validation_rules}
        self.input_order = [rule["name"] for rule in self.validation_rules]

        self._build_input_fields(self.headers)
        self._clear_treeview()

        cols = [f"c{i}" for i in range(len(self.headers))]
        self.tree["columns"] = cols
        for i, h in enumerate(self.headers):
            self.tree.heading(cols[i], text=h, anchor=tk.W)
            self.tree.column(cols[i], width=160, anchor=tk.W)

        start_row = header_row_idx + 1
        self.shadow.clear_sheet(self.current_sheet)
        rows = []

        for excel_row in range(start_row, sheet.max_row + 1):
            row_cells = [sheet.cell(row=excel_row, column=c + 1) for c in range(len(self.headers))]
            rowvals = [cell.value if cell.value is not None else "" for cell in row_cells]
            if all(v == "" or v is None for v in rowvals):
                continue

            display_row_index = excel_row - start_row + 1
            for col_idx, cell in enumerate(row_cells, start=1):
                self.shadow.set(self.current_sheet, display_row_index, col_idx,
                                cell.value, cell.value, False)
            rows.append(rowvals)

        self.all_rows = rows
        for r_idx, row in enumerate(rows, start=1):
            display_row = []
            for col_idx, cell_val in enumerate(row):
                header_name = self.input_order[col_idx]
                rule = self.current_rules.get(header_name, {})
                display_row.append(format_value_for_display(cell_val, rule=rule))
            row_id = self.tree.insert("", tk.END, values=display_row)
            self._apply_rounded_tags(row_id, display_row)

        self._create_filter_row()
        self._load_user_prefs()
        self._update_status(
            f"Loaded '{self.active_sheet_name}' — {len(rows)} rows, {len(self.headers)} columns.",
            "success"
        )

    # -----------------------------------------------------------------------
    # Validation rule UI
    # -----------------------------------------------------------------------

    def _clear_inputs_area(self):
        for w in self.inputs_inner.winfo_children():
            w.destroy()
        self.input_entries.clear()

    def _build_input_fields(self, headers):
        self._clear_inputs_area()
        for idx, header in enumerate(headers):
            rule = self.validation_rules[idx]
            col_frame = ttk.Frame(self.inputs_inner)
            col_frame.grid(row=0, column=idx, padx=6, pady=4)

            lbl = ttk.Label(col_frame, text=header, width=20, anchor="center")
            lbl.pack(side=tk.TOP, fill=tk.X)

            ent = tk.Entry(col_frame, width=20)
            ent.pack(side=tk.TOP, pady=(6, 0))
            ent.bind("<Return>", lambda e, i=idx: self._on_enter_pressed(e, i))
            ent.bind("<Tab>", lambda e, i=idx: (self._on_enter_pressed(e, i), "break")[1])
            self.input_entries.append(ent)

            error_var = tk.StringVar(value="")
            error_lbl = ttk.Label(col_frame, textvariable=error_var,
                                   foreground="red", anchor="center")
            error_lbl.pack(side=tk.TOP, fill=tk.X)
            ent.error_var = error_var

            control_frame = ttk.Frame(col_frame)
            control_frame.pack(side=tk.TOP, fill=tk.X, pady=(5, 0))

            req_chk = ttk.Checkbutton(control_frame, text="Required",
                                       variable=rule["required_var"],
                                       command=lambda r=rule: self._update_validation_state(r))
            req_chk.pack(anchor=tk.W)

            ttk.Label(control_frame, text="Duplicate Policy:").pack(anchor=tk.W, pady=(2, 0))
            for val, label in [("none", "None"), ("warn", "Warn"), ("strict", "Strict")]:
                ttk.Radiobutton(
                    control_frame, text=label,
                    variable=rule["duplicate_var"], value=val,
                    command=lambda r=rule: self._update_validation_state(r)
                ).pack(anchor=tk.W, padx=10)

        self.root.after(100, lambda: self.input_canvas.configure(
            scrollregion=self.input_canvas.bbox("all")
        ))
        self.reset_to_add_mode()

    def _update_validation_state(self, rule):
        rule["required"] = rule["required_var"].get()
        rule["duplicate_policy"] = rule["duplicate_var"].get()
        self._update_status(f"Updated policy for '{rule['name']}'")

    def _on_enter_pressed(self, event, idx: int):
        if idx == len(self.input_entries) - 1:
            if self.mode == "add":
                self.add_row_from_inputs()
            else:
                self.update_row_from_inputs()
        else:
            self.input_entries[idx + 1].focus_set()

    def _get_existing_column_data(self, col_index: int) -> set:
        if not self.workbook or not self.active_sheet_name:
            return set()
        sheet = self.workbook[self.active_sheet_name]
        result = set()
        for row_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=row_idx, column=col_index).value
            if val is not None:
                s = str(val).strip()
                if s:
                    result.add(s.lower())
        return result

    # -----------------------------------------------------------------------
    # Validation
    # -----------------------------------------------------------------------

    def validate_inputs(self):
        normalized = []
        strict_messages = []
        warning_messages = []
        is_valid = True
        is_edit_mode = self.mode == "edit"

        for i, entry in enumerate(self.input_entries):
            val = entry.get()
            rule = self.validation_rules[i]
            col_name = rule["name"]
            val_type = rule["type"]
            val_format = rule.get("format")
            required = rule["required"]
            duplicate_policy = rule["duplicate_policy"]

            try:
                entry.config(bg="white")
            except Exception as e:
                _log("validate/reset_bg", e)
            if hasattr(entry, "error_var"):
                entry.error_var.set("")

            val_stripped = str(val).strip()

            # 1. Required
            if required and not val_stripped:
                is_valid = False
                strict_messages.append(f"❌ {col_name}: Required.")
                normalized.append(None)
                entry.config(bg="#fbb")
                if hasattr(entry, "error_var"):
                    entry.error_var.set("Required")
                continue

            if not val_stripped and not required:
                normalized.append(None)
                continue

            # 2. Duplicate check
            if duplicate_policy in ("strict", "warn"):
                col_index = i + 1
                existing_data = self._get_existing_column_data(col_index=col_index)
                is_original_value = (
                    is_edit_mode
                    and val_stripped.lower() == self.original_editing_values.get(col_index, "__NO_MATCH__")
                )
                if not is_original_value and val_stripped.lower() in existing_data:
                    if duplicate_policy == "strict":
                        is_valid = False
                        strict_messages.append(f"❌ {col_name}: Duplicate value (Strict).")
                        entry.config(bg="#fbb")
                        if hasattr(entry, "error_var"):
                            entry.error_var.set("Duplicate")
                        normalized.append(val_stripped)
                        continue
                    else:
                        warning_messages.append(f"⚠️ {col_name}: Possible duplicate.")
                        entry.config(bg="#ffdd99")
                        if hasattr(entry, "error_var"):
                            entry.error_var.set("Duplicate?")

            # 3. Type parsing
            try:
                if val_type == "text":
                    cleaned = re.sub(r"\s+", " ", val_stripped)
                    normalized.append(cleaned if cleaned else None)

                elif val_type == "numeric":
                    if not is_numeric(val):
                        raise ValueError("Invalid number")
                    if val_format == "decimal" and has_excess_precision(val, decimal_limit=2):
                        warning_messages.append(
                            f"⚠️ {col_name}: Excess precision (>2 decimals). Will be rounded."
                        )
                        entry.config(bg="#e0ccff")
                        if hasattr(entry, "error_var"):
                            entry.error_var.set("Rounding alert")
                    normalized.append(normalize_numeric(val, fmt=val_format))

                elif val_type == "date":
                    date_obj = try_parse_date(val)
                    if date_obj is None:
                        raise ValueError("Invalid date")
                    normalized.append(date_obj)

                elif val_type == "email":
                    if not EMAIL_RE.match(val_stripped):
                        raise ValueError("Invalid email")
                    normalized.append(val_stripped)

                else:
                    normalized.append(val_stripped)

            except ValueError as e:
                is_valid = False
                strict_messages.append(f"❌ {col_name}: {e}")
                entry.config(bg="#fbb")
                if hasattr(entry, "error_var"):
                    entry.error_var.set(str(e))
                normalized.append(val_stripped)

        return is_valid, strict_messages, warning_messages, normalized

    def clear_input_entries(self):
        for ent in self.input_entries:
            try:
                ent.delete(0, tk.END)
                ent.config(bg="white")
            except Exception as e:
                _log("clear_input_entries", e)
            if hasattr(ent, "error_var"):
                ent.error_var.set("")

    # -----------------------------------------------------------------------
    # Rounded-value visual tags
    # -----------------------------------------------------------------------

    def _apply_rounded_tags(self, row_id, display_row):
        try:
            row_index_display = int(self.tree.index(row_id)) + 1
        except Exception as e:
            _log("_apply_rounded_tags/index", e)
            return

        for col_idx, text in enumerate(display_row):
            col_excel = col_idx + 1
            is_rounded = self.shadow.is_rounded(self.current_sheet, row_index_display, col_excel)
            if is_rounded:
                if self.rounded_icon:
                    try:
                        self.tree.image_create(row_id, column=col_excel - 1,
                                               image=self.rounded_icon, sticky="nw")
                    except Exception as e:
                        _log("_apply_rounded_tags/image", e)
                new_text = f"▲ {text}" if not str(text).startswith("▲ ") else text
            else:
                new_text = str(text)[2:] if str(text).startswith("▲ ") else text

            try:
                self.tree.set(row_id, col_idx, new_text)
            except Exception:
                vals = list(self.tree.item(row_id, "values"))
                vals[col_idx] = new_text
                self.tree.item(row_id, values=vals)

    # -----------------------------------------------------------------------
    # Add / Edit / Delete / Duplicate / Insert
    # -----------------------------------------------------------------------

    def add_row_from_inputs(self):
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        is_valid, strict_msgs, warn_msgs, normalized = self.validate_inputs()
        if not is_valid:
            self._update_status(f"Validation failed on {len(strict_msgs)} field(s).", "error")
            return

        if warn_msgs:
            res = messagebox.askyesno(
                "Input Warning",
                "Notices regarding your input:\n\n" + "\n".join(warn_msgs) + "\n\nProceed?",
                icon="warning"
            )
            if not res:
                self._update_status("Cancelled.", "warning")
                return

        sheet = self.workbook[self.active_sheet_name]
        append_row_excel = sheet.max_row + 1
        display_row = []
        display_row_index = len(self.tree.get_children()) + 1

        for idx, entry in enumerate(self.input_entries):
            raw = entry.get()
            parsed = normalized[idx]
            field_name = self.input_order[idx]
            rule = self.current_rules.get(field_name, {})
            display_text = format_value_for_display(parsed, rule=rule)

            rounded_flag = (
                parsed is not None
                and rule.get("format") == "decimal"
                and detect_precision_mismatch(raw, parsed, decimal_places=2)
            )

            col_index_excel = idx + 1
            self.shadow.set(self.current_sheet, display_row_index, col_index_excel,
                            raw, parsed, rounded_flag)
            display_row.append(display_text)

            # FIX: removed duplicate try/except where both branches did the same thing.
            try:
                sheet.cell(row=append_row_excel, column=col_index_excel).value = parsed
            except Exception as e:
                _log(f"add_row/cell_write col={col_index_excel}", e)
                self._update_status(f"Warning: could not write column {col_index_excel}.", "warning")

        row_id = self.tree.insert("", tk.END, values=display_row)
        self._apply_rounded_tags(row_id, display_row)
        self.all_rows.append(tuple(normalized))

        self.tree.selection_set(row_id)
        self.tree.see(row_id)
        self.unsaved_changes = True
        self._update_status(
            f"Added row {display_row_index} to '{self.active_sheet_name}'.", "success"
        )

        if self.auto_save_var.get():
            self.save_file()

        self.clear_input_entries()
        if self.input_entries:
            self.input_entries[0].focus_set()

    def update_row_from_inputs(self):
        if not self.editing_item:
            messagebox.showinfo("No selection", "No row selected for editing.")
            return

        is_valid, strict_msgs, warn_msgs, normalized = self.validate_inputs()
        if not is_valid:
            self._update_status("Validation failed.", "error")
            return

        if warn_msgs:
            res = messagebox.askyesno(
                "Update Warning",
                "Notices:\n\n" + "\n".join(warn_msgs) + "\n\nUpdate anyway?",
                icon="warning"
            )
            if not res:
                return

        display_row_index = int(self.tree.index(self.editing_item)) + 1
        sheet = self.workbook[self.active_sheet_name]
        excel_row = display_row_index + 1
        display_row = []

        for idx, entry in enumerate(self.input_entries):
            raw = entry.get()
            parsed = normalized[idx]
            field_name = self.input_order[idx]
            rule = self.current_rules.get(field_name, {})
            display_text = format_value_for_display(parsed, rule=rule)

            rounded_flag = (
                parsed is not None
                and rule.get("format") == "decimal"
                and detect_precision_mismatch(raw, parsed, decimal_places=2)
            )

            col_excel = idx + 1
            self.shadow.set(self.current_sheet, display_row_index, col_excel,
                            raw, parsed, rounded_flag)
            display_row.append(display_text)

            try:
                sheet.cell(row=excel_row, column=col_excel).value = parsed
            except Exception as e:
                _log(f"update_row/cell_write col={col_excel}", e)

        self.tree.item(self.editing_item, values=display_row)
        self._apply_rounded_tags(self.editing_item, display_row)
        self.unsaved_changes = True
        self._update_status(f"Updated row {display_row_index}.", "success")
        self.reset_to_add_mode()

        if self.auto_save_var.get():
            self.save_file()

    def on_tree_double_click(self, event):
        selected_item = self.tree.focus()
        if not selected_item:
            return
        values = self.tree.item(selected_item, "values")
        if not values:
            return

        self.clear_input_entries()
        display_row_index = int(self.tree.index(selected_item)) + 1

        for idx, entry in enumerate(self.input_entries):
            col_idx = idx + 1
            shadow_entry = self.shadow.get(self.current_sheet, display_row_index, col_idx)
            entry.delete(0, tk.END)
            if shadow_entry and shadow_entry.get("raw") is not None:
                entry.insert(0, shadow_entry.get("raw"))
            else:
                disp = values[idx] if idx < len(values) else ""
                if isinstance(disp, str) and disp.startswith("▲ "):
                    disp = disp[2:]
                entry.insert(0, disp)

        self.editing_item = selected_item
        self.mode = "edit"
        self.original_editing_values = {
            i + 1: str(entry.get()).strip().lower()
            for i, entry in enumerate(self.input_entries)
        }

        self.add_button.config(text="Update Row", command=self.update_row_from_inputs,
                               style="warning.TButton")
        for entry in self.input_entries:
            entry.unbind("<Return>")
            entry.bind("<Return>", lambda e: self.update_row_from_inputs())
        if self.input_entries:
            self.input_entries[0].focus_set()
        self._update_status("Editing existing row…", "warning")

    def edit_selected_row(self):
        sel = self.selected_item or self.tree.focus()
        if sel:
            self.tree.selection_set(sel)
            self.on_tree_double_click(None)

    def reset_to_add_mode(self):
        self.clear_input_entries()
        self.add_button.config(text="➕ Add Row", command=self.add_row_from_inputs,
                               style="success.TButton")
        for idx, entry in enumerate(self.input_entries):
            try:
                entry.unbind("<Return>")
                entry.unbind("<Tab>")
                entry.bind("<Return>", lambda e, i=idx: self._on_enter_pressed(e, i))
                entry.bind("<Tab>", lambda e, i=idx: (self._on_enter_pressed(e, i), "break")[1])
            except Exception as e:
                _log(f"reset_to_add_mode/bind idx={idx}", e)
        self.mode = "add"
        self.editing_item = None
        self.original_editing_values = {}
        if self.input_entries:
            self.input_entries[0].focus_set()

    def delete_selected_row(self):
        if not self.workbook or not self.active_sheet_name:
            return
        selected_item = self.tree.focus()
        if not selected_item:
            return

        confirm = messagebox.askyesno("Confirm Deletion", "Delete this row?")
        if not confirm:
            return

        try:
            self._flash_tree_row(selected_item, color="#ffcccc", duration=300)
            self.root.after(300, lambda: self._delete_row_after_flash(selected_item))
        except Exception as e:
            _log("delete_selected_row/flash", e)
            self._delete_row_after_flash(selected_item)

    def _flash_tree_row(self, item_id, color="#ccffcc", duration=800):
        try:
            tag_name = f"flash_{item_id}"
            self.tree.tag_configure(tag_name, background=color)
            self.tree.item(item_id, tags=(tag_name,))
            self.root.after(duration, lambda: self.tree.item(item_id, tags=()))
        except Exception as e:
            _log("_flash_tree_row", e)

    def _delete_row_after_flash(self, selected_item):
        try:
            display_row_index = int(self.tree.index(selected_item)) + 1
            sheet = self.workbook[self.active_sheet_name]
            excel_row_index = display_row_index + 1  # +1 for header row

            sheet.delete_rows(excel_row_index, 1)

            # FIX: re-index shadow store so rows below the deleted one stay correct.
            self.shadow.delete_row(self.current_sheet, display_row_index)

            # Keep all_rows cache in sync
            cache_idx = display_row_index - 1
            if 0 <= cache_idx < len(self.all_rows):
                self.all_rows.pop(cache_idx)

            self.tree.delete(selected_item)
            self.unsaved_changes = True
            total = len(self.tree.get_children())
            self._update_status(f"Row deleted. {total} row(s) remaining.", "success")

            if self.auto_save_var.get():
                self.save_file()

        except Exception as e:
            _log("_delete_row_after_flash", e)
            messagebox.showerror("Error", f"Failed to delete row:\n{e}")

    # FIX: Ctrl+D — duplicate row into input fields (was documented, never coded).
    def duplicate_selected_row(self):
        sel = self.selected_item or self.tree.focus()
        if not sel:
            return
        values = self.tree.item(sel, "values")
        if not values:
            return

        display_row_index = int(self.tree.index(sel)) + 1
        self.clear_input_entries()
        self.reset_to_add_mode()

        for idx, entry in enumerate(self.input_entries):
            col_idx = idx + 1
            shadow_entry = self.shadow.get(self.current_sheet, display_row_index, col_idx)
            raw = shadow_entry.get("raw") if shadow_entry and shadow_entry.get("raw") else ""
            if not raw and idx < len(values):
                raw = values[idx]
                if isinstance(raw, str) and raw.startswith("▲ "):
                    raw = raw[2:]
            entry.insert(0, raw)

        self._update_status("Row duplicated into input fields — review and press Enter/Add to save.", "info")
        if self.input_entries:
            self.input_entries[0].focus_set()

    # FIX: Ctrl+Shift+I — insert blank row at selection (was documented, never coded).
    def insert_blank_row(self):
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        sel = self.selected_item or self.tree.focus()
        if not sel:
            # No selection — append at end
            self.clear_input_entries()
            self._update_status("Blank row ready — fill in and press Add.", "info")
            if self.input_entries:
                self.input_entries[0].focus_set()
            return

        display_row_index = int(self.tree.index(sel)) + 1
        excel_row_index = display_row_index + 1  # +1 for header

        sheet = self.workbook[self.active_sheet_name]
        sheet.insert_rows(excel_row_index)

        # Shift shadow store entries below insertion point down by 1
        new_data = {}
        for (sname, ridx, cidx), val in self.shadow._data.items():
            if sname == self.current_sheet and ridx >= display_row_index:
                new_data[(sname, ridx + 1, cidx)] = val
            else:
                new_data[(sname, ridx, cidx)] = val
        self.shadow._data = new_data

        # Insert blank into treeview and cache
        blank = [""] * len(self.headers)
        self.tree.insert("", display_row_index - 1, values=blank)
        self.all_rows.insert(display_row_index - 1, tuple(blank))

        self.unsaved_changes = True
        self._update_status(f"Blank row inserted at position {display_row_index}.", "success")

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------

    def save_file(self) -> bool:
        if not self.workbook:
            return False
        if not self.filepath:
            return self.save_file_as()
        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved: {os.path.basename(self.filepath)}", "success")
            self._save_user_prefs()
            return True
        except Exception as e:
            messagebox.showerror("Save error", str(e))
            return False

    def save_file_as(self) -> bool:
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return False
        self.filepath = path
        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved as: {os.path.basename(self.filepath)}", "success")
            self._save_user_prefs()
            return True
        except Exception as e:
            messagebox.showerror("Save error", str(e))
            return False

    # -----------------------------------------------------------------------
    # Preferences
    # -----------------------------------------------------------------------

    def _load_user_prefs(self):
        prefs = load_user_prefs(self.filepath)
        if not prefs:
            return
        try:
            if "theme" in prefs:
                self.theme_combo.set(prefs["theme"])
                Style().theme_use(prefs["theme"])
        except Exception as e:
            _log("_load_user_prefs/theme", e)
        try:
            if "auto_save" in prefs:
                self.auto_save_var.set(bool(prefs["auto_save"]))
        except Exception as e:
            _log("_load_user_prefs/auto_save", e)

        sheet_prefs = prefs.get("sheets", {}).get(self.active_sheet_name, {})
        for rule in self.validation_rules:
            col_pref = sheet_prefs.get("columns", {}).get(rule["name"])
            if col_pref:
                try:
                    rule["required_var"].set(bool(col_pref.get("required")))
                    rule["duplicate_var"].set(col_pref.get("duplicate", "none"))
                    self._update_validation_state(rule)
                except Exception as e:
                    _log(f"_load_user_prefs/col={rule['name']}", e)

    def _save_user_prefs(self):
        prefs = {
            "theme": self.theme_combo.get(),
            "auto_save": bool(self.auto_save_var.get()),
            "sheets": {
                self.active_sheet_name: {
                    "columns": {
                        rule["name"]: {
                            "required": bool(rule["required_var"].get()),
                            "duplicate": rule["duplicate_var"].get(),
                        }
                        for rule in self.validation_rules
                    }
                }
            },
        }
        save_user_prefs(self.filepath, prefs)

    # -----------------------------------------------------------------------
    # Misc UI helpers
    # -----------------------------------------------------------------------

    def on_theme_change(self, event=None):
        try:
            Style().theme_use(self.theme_combo.get())
            self._update_status(f"Theme: {self.theme_combo.get()}", "success")
        except Exception as e:
            _log("on_theme_change", e)

    def _add_hover_effect(self, widget):
        widget.bind("<Enter>", lambda e: widget.configure(cursor="hand2"))
        widget.bind("<Leave>", lambda e: widget.configure(cursor=""))

    def _show_about(self):
        messagebox.showinfo("About", APP_TITLE)

    def _show_help(self):
        """
        FIX: was Windows-only (os.startfile). Now works on macOS and Linux too.
        """
        help_path = resource_path("help.txt")
        if os.path.exists(help_path):
            try:
                open_file_cross_platform(help_path)
            except Exception as e:
                messagebox.showwarning("Help", f"Could not open help file:\n{e}")
        else:
            messagebox.showinfo("Help", "No help.txt found next to the application.")

    def on_close(self):
        if self.unsaved_changes:
            if messagebox.askyesno("Unsaved changes", "Save before exit?"):
                self.save_file()
        self._save_user_prefs()
        self.root.destroy()
